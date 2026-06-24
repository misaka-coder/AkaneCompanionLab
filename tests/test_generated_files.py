from __future__ import annotations

import json
import subprocess
import tempfile
import unittest
import wave
import zipfile
from array import array
from pathlib import Path
from unittest.mock import patch

from companion_v01.attachment_inbox import AttachmentInboxService
from companion_v01.generated_files import GeneratedFileService
from companion_v01.store import MemoryStore
from companion_v01.tool_runtime import ApplyStyleToExistingFileToolHandler, CleanVoiceTrackToolHandler, ComposeFileToolHandler, ConvertMediaFileToolHandler, InspectGeneratedFileToolHandler, InspectMediaInfoToolHandler, ManageGeneratedFileToolHandler, PrepareVoiceDatasetToolHandler, ReviseGeneratedFileToolHandler, SendFileToolHandler, SendGeneratedFileToolHandler, SeparateAudioStemsToolHandler, ToolExecutionContext, TranscribeMediaToolHandler


def _write_test_wav(path: Path, *, sample_rate: int = 1000) -> None:
    samples = array("h")

    def extend_silence(seconds: float) -> None:
        samples.extend([0] * int(sample_rate * seconds))

    def extend_tone(seconds: float, amplitude: int = 9000) -> None:
        count = int(sample_rate * seconds)
        for index in range(count):
            samples.append(amplitude if (index // 25) % 2 == 0 else -amplitude)

    extend_silence(0.4)
    extend_tone(4.0)
    extend_silence(0.6)
    extend_tone(1.0)
    extend_silence(0.6)
    extend_tone(14.0)
    extend_silence(0.4)
    path.parent.mkdir(parents=True, exist_ok=True)
    with wave.open(str(path), "wb") as writer:
        writer.setnchannels(1)
        writer.setsampwidth(2)
        writer.setframerate(sample_rate)
        writer.writeframes(samples.tobytes())


class GeneratedFileTests(unittest.TestCase):
    def test_store_generated_file_roundtrip_and_handles(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            store = MemoryStore(Path(temp_dir))
            first = store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                output_title="学习计划",
                output_format="md",
                storage_relpath="user/session/plan.md",
                source_ids=["file_001"],
                content_card={"summary": "学习计划摘要"},
                summary="学习计划摘要",
                timestamp=100,
            )
            second = store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                output_title="学习计划 Word",
                output_format="docx",
                storage_relpath="user/session/plan.docx",
                timestamp=110,
            )

            self.assertEqual(first["generated_handle"], "gen_001")
            self.assertEqual(second["generated_handle"], "gen_002")
            self.assertEqual(first["source_ids"], ["file_001"])
            self.assertEqual(first["content_card"]["summary"], "学习计划摘要")

            listed = store.list_generated_files(
                profile_user_id="user",
                session_id="session",
                statuses=["ready"],
                limit=10,
            )
            self.assertEqual([item["generated_handle"] for item in listed], ["gen_002", "gen_001"])

            found = store.find_generated_file(
                profile_user_id="user",
                session_id="session",
                query="学习计划",
            )
            self.assertIsNotNone(found)
            self.assertIn(found["generated_handle"], {"gen_001", "gen_002"})

    def test_find_generated_file_prefers_exact_title_over_newer_partial_match(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            store = MemoryStore(Path(temp_dir))
            vocal = store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                output_title="昔涟_人声",
                output_format="flac",
                storage_relpath="user/session/vocal.flac",
                timestamp=100,
            )
            denoised = store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                output_title="昔涟_人声_降噪",
                output_format="flac",
                storage_relpath="user/session/vocal_denoised.flac",
                timestamp=200,
            )

            found = store.find_generated_file(
                profile_user_id="user",
                session_id="session",
                query="昔涟_人声",
                statuses=["ready"],
            )

            self.assertIsNotNone(found)
            self.assertEqual(found["generated_id"], vocal["generated_id"])
            self.assertNotEqual(found["generated_id"], denoised["generated_id"])

    def test_find_generated_file_prefers_exact_handle_over_similar_title(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            store = MemoryStore(Path(temp_dir))
            target = store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                generated_handle="gen_033",
                output_title="原始人声",
                output_format="flac",
                storage_relpath="user/session/vocal.flac",
                timestamp=100,
            )
            store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                generated_handle="gen_035",
                output_title="这是 gen_033 的降噪版",
                output_format="flac",
                storage_relpath="user/session/vocal_denoised.flac",
                timestamp=200,
            )

            found = store.find_generated_file(
                profile_user_id="user",
                session_id="session",
                query="gen_033",
                statuses=["ready"],
            )

            self.assertIsNotNone(found)
            self.assertEqual(found["generated_id"], target["generated_id"])
            self.assertEqual(found["generated_handle"], "gen_033")

    def test_find_generated_file_requires_confirmation_for_ambiguous_fuzzy_match(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            store = MemoryStore(Path(temp_dir))
            store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                output_title="昔涟_人声",
                output_format="flac",
                storage_relpath="user/session/vocal.flac",
                timestamp=100,
            )
            store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                output_title="昔涟_人声_降噪",
                output_format="flac",
                storage_relpath="user/session/vocal_denoised.flac",
                timestamp=200,
            )

            found = store.find_generated_file(
                profile_user_id="user",
                session_id="session",
                query="人声",
                statuses=["ready"],
            )

            self.assertIsNone(found)

    def test_generated_prompt_context_exposes_media_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                output_title="昔涟_人声",
                output_format="flac",
                storage_relpath="user/session/vocal.flac",
                file_size=17425491,
                content_card={
                    "summary": "从 file_016 分离出人声轨。",
                    "source": {"handle": "file_016", "title": "昔涟"},
                    "separation": {"stem_role": "vocals"},
                    "media_info": {
                        "format_name": "flac",
                        "duration_seconds": 192.0,
                        "file_size": 17425491,
                        "audio": {
                            "codec": "flac",
                            "sample_rate": 48000,
                            "channels": 2,
                            "bit_rate": 1411000,
                        },
                    },
                },
                summary="从 file_016 分离出人声轨。",
                created_by_tool="separate_audio_stems",
                timestamp=100,
            )

            context = generated_service.build_prompt_context(
                profile_user_id="user",
                session_id="session",
                limit=3,
            )

            self.assertIn("大小：", context)
            self.assertIn("来源工具：separate_audio_stems", context)
            self.assertIn("来源：file_016", context)
            self.assertIn("音轨角色：vocals", context)
            self.assertIn("媒体规格：格式：flac", context)
            self.assertIn("时长：3:12", context)
            self.assertIn("音频：编码 flac，48000Hz，2声道，1.41Mbps。", context)

    def test_compose_file_creates_markdown_from_attachment(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="document",
                origin_name="notes.txt",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="课堂笔记",
                short_hint="讲了向量检索和工具调用。",
                detail={"text_preview": "RAG 可以把长期记忆和当前上下文结合起来。"},
                timestamp=110,
            )

            result = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=["file_001"],
                task="整理成学习笔记",
                output_format="md",
                output_title="RAG 学习笔记",
                content_markdown="# RAG 学习笔记\n\n- 向量检索\n- 工具调用",
                timestamp=120,
            )

            self.assertTrue(result["ok"])
            generated = result["generated"]
            self.assertEqual(generated["generated_handle"], "gen_001")
            output_path = Path(generated["absolute_path"])
            self.assertTrue(output_path.exists())
            self.assertIn("向量检索", output_path.read_text(encoding="utf-8"))

            prompt = generated_service.build_prompt_context(
                profile_user_id="user",
                session_id="session",
            )
            self.assertIn("生成文件工作台", prompt)
            self.assertIn("gen_001", prompt)
            self.assertIn("RAG 学习笔记", prompt)

    def test_compose_file_fallback_uses_larger_original_attachment_material(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "long.txt"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_text("\n".join(f"完整第{index}行" for index in range(1, 260)), encoding="utf-8")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="document",
                origin_name="long.txt",
                storage_relpath="master/long.txt",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="长文本",
                short_hint="预览只有开头。",
                detail={"text_preview": "完整第1行\n完整第2行", "preview_is_truncated": True},
                timestamp=110,
            )

            result = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=["file_001"],
                task="转成 Markdown",
                output_format="md",
                output_title="长文本整理",
                timestamp=120,
            )

            self.assertTrue(result["ok"])
            output_text = Path(result["generated"]["absolute_path"]).read_text(encoding="utf-8")
            self.assertIn("完整第200行", output_text)
            self.assertNotIn("任务：", output_text)
            self.assertNotIn("来源摘录", output_text)
            self.assertNotIn("注意：上面是系统可安全展开的片段", output_text)

    def test_compose_file_faithful_conversion_recovers_from_prompt_excerpt(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            full_text = "\n".join(
                [
                    "source plugin:https://example.com/index.json",
                    "TachiyomiJ2K:https://example.com/tachiyomi",
                    "Mihon:https://example.com/mihon",
                    "komikku:https://example.com/komikku",
                    "tail:https://example.com/extensions-source",
                ]
            )
            stored = attachment_root / "master" / "links.txt"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_text(full_text, encoding="utf-8")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="document",
                origin_name="links.txt",
                storage_relpath="master/links.txt",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="链接列表",
                short_hint="一些链接。",
                detail={"text_preview": full_text},
                timestamp=110,
            )

            result = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=["file_001"],
                task="忠实转换为 PDF，保留原文内容",
                output_format="txt",
                output_title="链接列表转换",
                content_markdown=full_text[:120],
                send_to_user=False,
                timestamp=120,
            )

            self.assertTrue(result["ok"])
            output_text = Path(result["generated"]["absolute_path"]).read_text(encoding="utf-8")
            self.assertIn("tail:https://example.com/extensions-source", output_text)
            self.assertNotIn("任务：", output_text)
            self.assertNotIn("来源摘录", output_text)

    def test_compose_file_faithful_generated_source_does_not_include_source_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            original = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一封心里话",
                output_format="md",
                output_title="我想对主人说的话",
                content_markdown="亲爱的主人：\n\n谢谢你一直陪着我。\n\n最喜欢主人的 Akane",
                send_to_user=False,
                timestamp=100,
            )
            self.assertTrue(original["ok"])

            converted = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=["gen_001"],
                task="将刚刚写给主人的心里话文档转换成 PDF 格式",
                output_format="txt",
                output_title="我想对主人说的话",
                send_to_user=False,
                timestamp=110,
            )

            self.assertTrue(converted["ok"])
            output_text = Path(converted["generated"]["absolute_path"]).read_text(encoding="utf-8")
            self.assertIn("亲爱的主人", output_text)
            self.assertIn("最喜欢主人的 Akane", output_text)
            self.assertNotIn("任务：", output_text)
            self.assertNotIn("来源摘录", output_text)
            self.assertNotIn("用途：", output_text)

    def test_compose_file_summary_does_not_replace_with_full_source_prefix(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            full_text = "第一段\n第二段\n第三段\n第四段"
            stored = attachment_root / "master" / "notes.txt"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_text(full_text, encoding="utf-8")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="document",
                origin_name="notes.txt",
                storage_relpath="master/notes.txt",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="笔记",
                short_hint="一些段落。",
                detail={"text_preview": full_text},
                timestamp=110,
            )

            result = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=["file_001"],
                task="摘要成一个短文件",
                output_format="txt",
                output_title="摘要",
                content_markdown="第一段",
                send_to_user=False,
                timestamp=120,
            )

            self.assertTrue(result["ok"])
            output_text = Path(result["generated"]["absolute_path"]).read_text(encoding="utf-8")
            self.assertEqual(output_text, "第一段")

    def test_compose_file_tool_handler_emits_generated_event(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            handler = ComposeFileToolHandler(generated_file_service=generated_service)
            call = handler.normalize_call(
                {
                    "type": "compose_file",
                    "output_format": "txt",
                    "output_title": "小结",
                    "content_markdown": "这是整理好的内容。",
                    "formatting": {
                        "header": {"bold": True},
                        "columns": [{"match_header": "姓名", "font_color": "red", "unknown": "ignored"}],
                        "unsafe": "ignored",
                    },
                    "send_to_user": True,
                }
            )
            self.assertIsNotNone(call)
            self.assertIn("formatting", call or {})
            self.assertNotIn("unsafe", (call or {}).get("formatting", {}))
            self.assertNotIn("unknown", (call or {}).get("formatting", {}).get("columns", [{}])[0])
            context = ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=123,
                visual_payload={},
            )
            result = handler.execute(call=call or {}, context=context)

            self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
            self.assertIn("gen_001", result.followup_context)
            self.assertTrue(result.stream_events[0]["send_to_user"])

    def test_compose_file_instruction_discourages_verbal_only_promises(self) -> None:
        handler = ComposeFileToolHandler(generated_file_service=object())
        instruction = handler.build_prompt_instruction()

        self.assertIn("开始/继续/直接做", instruction)
        self.assertIn("不要只口头答应", instruction)
        self.assertIn("tool_call 调用 compose_file", instruction)

    def test_convert_media_file_creates_generated_audio(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "song.flac"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_bytes(b"fake flac payload")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="audio",
                origin_name="song.flac",
                storage_relpath="master/song.flac",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="song.flac",
                short_hint="一首音频。",
                detail={"file_kind": "flac"},
                timestamp=110,
            )

            def fake_run(command, **kwargs):
                Path(command[-1]).write_bytes(b"fake mp3 payload")

                class Result:
                    returncode = 0
                    stdout = ""
                    stderr = ""

                return Result()

            with patch("companion_v01.generated_files.shutil.which", return_value="ffmpeg"), patch(
                "companion_v01.generated_files.subprocess.run",
                side_effect=fake_run,
            ) as mocked_run:
                result = generated_service.convert_media_file(
                    profile_user_id="user",
                    session_id="session",
                    source_target="audio_001",
                    output_format="mp3",
                    output_title="song_mp3",
                    bitrate="192k",
                    sample_rate=44100,
                    channels=2,
                    timestamp=120,
                )

            self.assertTrue(result["ok"])
            generated = result["generated"]
            self.assertEqual(generated["output_format"], "mp3")
            self.assertEqual(generated["file_ext"], "mp3")
            self.assertEqual(generated["mime_type"], "audio/mpeg")
            self.assertEqual(generated["created_by_tool"], "convert_media_file")
            self.assertTrue(Path(generated["absolute_path"]).exists())
            media_info = (generated.get("content_card") or {}).get("media_info", {})
            self.assertEqual(media_info.get("format_name"), "mp3")
            self.assertEqual((media_info.get("audio") or {}).get("sample_rate"), 44100)
            self.assertEqual((media_info.get("audio") or {}).get("channels"), 2)
            self.assertEqual((media_info.get("audio") or {}).get("bit_rate"), 192000)
            command = mocked_run.call_args.args[0]
            self.assertIn("-b:a", command)
            self.assertIn("192k", command)

    def test_convert_media_file_supports_trim_filters_and_speed(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "clip.mp4"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_bytes(b"fake video payload")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="file",
                origin_name="clip.mp4",
                storage_relpath="master/clip.mp4",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="clip.mp4",
                short_hint="一段视频。",
                detail={"file_kind": "mp4"},
                timestamp=110,
            )

            def fake_run(command, **kwargs):
                class Result:
                    returncode = 0
                    stdout = ""
                    stderr = ""

                if command[0] == "ffprobe":
                    Result.stdout = "120.0"
                    return Result()
                Path(command[-1]).write_bytes(b"fake mp3 payload")
                return Result()

            with patch("companion_v01.generated_files.shutil.which", side_effect=lambda name: name), patch(
                "companion_v01.generated_files.subprocess.run",
                side_effect=fake_run,
            ) as mocked_run:
                result = generated_service.convert_media_file(
                    profile_user_id="user",
                    session_id="session",
                    source_target="file_001",
                    output_format="mp3",
                    output_title="clip_audio",
                    start_time="00:00:10",
                    end_time="40s",
                    normalize_volume=True,
                    volume_gain_db=6,
                    trim_silence=True,
                    fade_in_seconds=2,
                    fade_out_seconds=3,
                    speed_ratio=1.5,
                    timestamp=120,
                )

            self.assertTrue(result["ok"])
            ffmpeg_command = mocked_run.call_args_list[-1].args[0]
            self.assertIn("-ss", ffmpeg_command)
            self.assertIn("10", ffmpeg_command)
            self.assertIn("-t", ffmpeg_command)
            self.assertIn("30", ffmpeg_command)
            filter_text = ffmpeg_command[ffmpeg_command.index("-filter:a") + 1]
            self.assertIn("silenceremove=", filter_text)
            self.assertIn("areverse", filter_text)
            self.assertIn("atempo=1.5", filter_text)
            self.assertIn("loudnorm", filter_text)
            self.assertIn("volume=6dB", filter_text)
            self.assertIn("afade=t=in:st=0:d=2", filter_text)
            self.assertIn("afade=t=out:st=17:d=3", filter_text)
            conversion = result["generated"]["content_card"]["conversion"]
            self.assertEqual(conversion["start_seconds"], 10.0)
            self.assertEqual(conversion["duration_seconds"], 30.0)
            self.assertTrue(conversion["normalize_volume"])
            self.assertEqual(conversion["volume_gain_db"], 6)
            self.assertTrue(conversion["trim_silence"])
            self.assertEqual(conversion["speed_ratio"], 1.5)

    def test_inspect_media_info_reads_ffprobe_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "clip.mp4"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_bytes(b"fake video payload")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="file",
                origin_name="clip.mp4",
                storage_relpath="master/clip.mp4",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="clip.mp4",
                short_hint="一段视频。",
                detail={"file_kind": "mp4"},
                timestamp=110,
            )

            def fake_run(command, **kwargs):
                class Result:
                    returncode = 0
                    stderr = ""
                    stdout = """
                    {
                      "format": {
                        "format_name": "mov,mp4,m4a,3gp,3g2,mj2",
                        "duration": "184.2",
                        "size": "1234567",
                        "bit_rate": "320000"
                      },
                      "streams": [
                        {
                          "index": 0,
                          "codec_type": "video",
                          "codec_name": "h264",
                          "width": 1920,
                          "height": 1080,
                          "avg_frame_rate": "30000/1001"
                        },
                        {
                          "index": 1,
                          "codec_type": "audio",
                          "codec_name": "aac",
                          "sample_rate": "44100",
                          "channels": 2,
                          "bit_rate": "128000"
                        }
                      ]
                    }
                    """

                return Result()

            with patch("companion_v01.generated_files.shutil.which", return_value="ffprobe"), patch(
                "companion_v01.generated_files.subprocess.run",
                side_effect=fake_run,
            ):
                result = generated_service.inspect_media_info(
                    profile_user_id="user",
                    session_id="session",
                    source_target="file_001",
                    timestamp=120,
                )

            self.assertTrue(result["ok"])
            media_info = result["media_info"]
            self.assertEqual(media_info["duration_seconds"], 184.2)
            self.assertEqual(media_info["audio"]["codec"], "aac")
            self.assertEqual(media_info["audio"]["sample_rate"], 44100)
            self.assertEqual(media_info["video"]["width"], 1920)
            self.assertEqual(media_info["video"]["height"], 1080)
            self.assertAlmostEqual(media_info["video"]["fps"], 29.97, places=2)
            self.assertIn("音频", result["followup_context"])
            self.assertIn("视频", result["followup_context"])

    def test_separate_audio_stems_creates_vocals_and_instrumental_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "song.wav"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_bytes(b"fake wav payload")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="audio",
                origin_name="song.wav",
                storage_relpath="master/song.wav",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="song.wav",
                short_hint="一首普通音频。",
                detail={"file_kind": "wav", "media_info": {"audio": {"codec": "pcm_s16le"}}},
                timestamp=110,
            )

            def fake_module_separation(self, *, source_path, output_root, model_name="htdemucs"):
                output_root.mkdir(parents=True, exist_ok=True)
                vocals = output_root / "vocals.wav"
                instrumental = output_root / "instrumental.wav"
                vocals.write_bytes(b"fake vocals")
                instrumental.write_bytes(b"fake accompaniment")
                return {
                    "vocals": vocals,
                    "instrumental": instrumental,
                }

            with patch("companion_v01.generated_files.importlib.util.find_spec", return_value=object()), patch.object(
                GeneratedFileService,
                "_separate_audio_with_demucs_module",
                new=fake_module_separation,
            ):
                result = generated_service.separate_audio_stems(
                    profile_user_id="user",
                    session_id="session",
                    source_target="audio_001",
                    mode="vocals_instrumental",
                    output_format="wav",
                    output_title="副歌练习分轨",
                    timestamp=120,
                )

            self.assertTrue(result["ok"])
            generated_files = result["generated_files"]
            self.assertEqual(len(generated_files), 2)
            self.assertEqual([item["generated_handle"] for item in generated_files], ["gen_001", "gen_002"])
            self.assertEqual(generated_files[0]["created_by_tool"], "separate_audio_stems")
            self.assertTrue(Path(generated_files[0]["absolute_path"]).exists())
            self.assertTrue(Path(generated_files[1]["absolute_path"]).exists())
            roles = [
                (item.get("content_card") or {}).get("separation", {}).get("stem_role")
                for item in generated_files
            ]
            self.assertEqual(roles, ["vocals", "instrumental"])
            self.assertEqual((generated_files[0].get("content_card") or {}).get("media_info", {}).get("format_name"), "wav")
            self.assertEqual((generated_files[1].get("content_card") or {}).get("media_info", {}).get("format_name"), "wav")
            self.assertIn("人声 / 伴奏分离", result["followup_context"])

    def test_separate_audio_stems_tool_handler_emits_two_generated_events(self) -> None:
        class FakeGeneratedService:
            def separate_audio_stems(self, **kwargs):
                return {
                    "ok": True,
                    "generated_files": [
                        {"generated_id": "generated::1", "generated_handle": "gen_001"},
                        {"generated_id": "generated::2", "generated_handle": "gen_002"},
                    ],
                    "send_to_user": True,
                    "followup_context": "分离完成。",
                }

        handler = SeparateAudioStemsToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call(
            {
                "type": "separate_audio_stems",
                "source_id": "audio_001",
                "mode": "人声伴奏",
                "output_format": "wave",
                "send_to_user": True,
            }
        )

        self.assertIsNotNone(call)
        self.assertEqual((call or {}).get("mode"), "vocals_instrumental")
        self.assertEqual((call or {}).get("output_format"), "wav")
        result = handler.execute(
            call=call or {},
            context=ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=100,
                visual_payload={},
            ),
        )

        self.assertEqual(len(result.stream_events), 2)
        self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
        self.assertEqual(result.stream_events[1]["generated_file"]["generated_handle"], "gen_002")
        self.assertEqual(result.followup_context, "分离完成。")

    def test_clean_voice_track_creates_generated_audio(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "voice.wav"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_bytes(b"fake wav payload")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="audio",
                origin_name="voice.wav",
                storage_relpath="master/voice.wav",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="voice.wav",
                short_hint="一段带点底噪的人声。",
                detail={"file_kind": "wav", "media_info": {"audio": {"codec": "pcm_s16le"}}},
                timestamp=110,
            )

            def fake_run(command, capture_output, text, timeout, check):
                if "-codec:a" in command and "pcm_s16le" in command:
                    Path(command[-1]).write_bytes(b"prepared wav")
                    return subprocess.CompletedProcess(command, 0, "", "")
                if command and str(command[0]).lower().startswith("deepfilter"):
                    out_dir = Path(command[command.index("-o") + 1])
                    out_dir.mkdir(parents=True, exist_ok=True)
                    (out_dir / "voice.wav").write_bytes(b"cleaned wav")
                    return subprocess.CompletedProcess(command, 0, "", "")
                raise AssertionError(f"unexpected command: {command}")

            with patch("companion_v01.generated_files.shutil.which", return_value="ffmpeg.exe"), patch.object(
                GeneratedFileService,
                "_resolve_deepfilternet_runner",
                return_value={"kind": "binary", "command": ["deepFilter.exe"]},
            ), patch("companion_v01.generated_files.subprocess.run", side_effect=fake_run):
                result = generated_service.clean_voice_track(
                    profile_user_id="user",
                    session_id="session",
                    source_target="audio_001",
                    mode="denoise",
                    quality="auto",
                    output_format="wav",
                    output_title="voice_clean",
                    timestamp=120,
                )

            self.assertTrue(result["ok"])
            generated = result["generated"]
            self.assertEqual(generated["generated_handle"], "gen_001")
            self.assertEqual(generated["created_by_tool"], "clean_voice_track")
            self.assertTrue(Path(generated["absolute_path"]).exists())
            self.assertEqual(
                (generated.get("content_card") or {}).get("voice_cleaning", {}).get("backend_used"),
                "deepfilternet",
            )
            media_info = (generated.get("content_card") or {}).get("media_info", {})
            self.assertEqual(media_info.get("format_name"), "wav")
            self.assertEqual((media_info.get("audio") or {}).get("sample_rate"), 48000)
            self.assertEqual((media_info.get("audio") or {}).get("channels"), 1)
            self.assertIn("AI 净化", result["followup_context"])

    def test_clean_voice_track_tool_handler_emits_generated_event(self) -> None:
        class FakeGeneratedService:
            def clean_voice_track(self, **kwargs):
                return {
                    "ok": True,
                    "generated": {"generated_id": "generated::1", "generated_handle": "gen_001"},
                    "send_to_user": True,
                    "followup_context": "净化完成。",
                }

        handler = CleanVoiceTrackToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call(
            {
                "type": "clean_voice_track",
                "source_id": "audio_001",
                "mode": "去混响",
                "quality": "deepfilternet",
                "output_format": "wave",
                "post_filter": True,
                "send_to_user": True,
            }
        )

        self.assertIsNotNone(call)
        self.assertEqual((call or {}).get("mode"), "dereverb")
        self.assertEqual((call or {}).get("quality"), "ai")
        self.assertEqual((call or {}).get("output_format"), "wav")
        result = handler.execute(
            call=call or {},
            context=ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=100,
                visual_payload={},
            ),
        )

        self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
        self.assertEqual(result.stream_events[0]["generated_file"]["generated_handle"], "gen_001")
        self.assertEqual(result.followup_context, "净化完成。")

    def test_prepare_voice_dataset_creates_zip_manifest_with_issue_filenames(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "voice.wav"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_bytes(b"fake wav payload")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="audio",
                origin_name="voice.wav",
                storage_relpath="master/voice.wav",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="voice.wav",
                short_hint="一段待切片的人声。",
                detail={"file_kind": "wav", "media_info": {"audio": {"codec": "pcm_s16le"}}},
                timestamp=110,
            )

            def fake_run(command, capture_output, text, timeout, check):
                _write_test_wav(Path(command[-1]), sample_rate=1000)
                return subprocess.CompletedProcess(command, 0, "", "")

            with patch("companion_v01.generated_files.shutil.which", return_value="ffmpeg.exe"), patch(
                "companion_v01.generated_files.subprocess.run",
                side_effect=fake_run,
            ):
                result = generated_service.prepare_voice_dataset(
                    profile_user_id="user",
                    session_id="session",
                    source_targets=["audio_001"],
                    profile="gpt_sovits",
                    output_title="akane_dataset",
                    target_sr=44100,
                    min_clip_seconds=3,
                    max_clip_seconds=12,
                    silence_threshold_db=-40,
                    min_silence_ms=300,
                    max_silence_kept_ms=100,
                    timestamp=120,
                )

            self.assertTrue(result["ok"])
            generated = result["generated"]
            self.assertEqual(generated["generated_handle"], "gen_001")
            self.assertEqual(generated["output_format"], "zip")
            self.assertEqual(generated["mime_type"], "application/zip")
            self.assertEqual(generated["created_by_tool"], "prepare_voice_dataset")
            output_path = Path(generated["absolute_path"])
            self.assertTrue(output_path.exists())
            with zipfile.ZipFile(output_path) as archive:
                names = archive.namelist()
                self.assertIn("manifest.json", names)
                self.assertTrue(any(name.startswith("slices/src01_slice_") for name in names))
                manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
            self.assertGreaterEqual(manifest["stats"]["slice_count"], 3)
            issue_slices = manifest["issue_slices"]
            self.assertIn("too_short", issue_slices)
            self.assertIn("too_long", issue_slices)
            card_preview = (generated.get("content_card") or {}).get("content_preview", "")
            self.assertIn("过短", card_preview)
            self.assertIn("过长", card_preview)
            self.assertIn("prepare_voice_dataset", result["followup_context"])

    def test_prepare_voice_dataset_tool_handler_emits_generated_event(self) -> None:
        class FakeGeneratedService:
            def prepare_voice_dataset(self, **kwargs):
                return {
                    "ok": True,
                    "generated": {"generated_id": "generated::1", "generated_handle": "gen_001"},
                    "send_to_user": True,
                    "followup_context": "训练集完成。",
                }

        handler = PrepareVoiceDatasetToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call(
            {
                "type": "prepare_voice_dataset",
                "source_ids": ["gen_001", "audio_002"],
                "profile": "sovits",
                "min_clip_seconds": 3,
                "max_clip_seconds": 12,
                "send_to_user": True,
            }
        )

        self.assertIsNotNone(call)
        self.assertEqual((call or {}).get("source_ids"), ["gen_001", "audio_002"])
        self.assertEqual((call or {}).get("profile"), "gpt_sovits")
        result = handler.execute(
            call=call or {},
            context=ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=100,
                visual_payload={},
            ),
        )

        self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
        self.assertEqual(result.stream_events[0]["generated_file"]["generated_handle"], "gen_001")
        self.assertEqual(result.followup_context, "训练集完成。")

    def test_transcribe_media_creates_merged_markdown_transcript(self) -> None:
        class FakeWhisperModel:
            def transcribe(self, audio_path, **kwargs):
                return (
                    [
                        {"start": 0.0, "end": 1.2, "text": "你好，主人。"},
                        {"start": 1.5, "end": 3.0, "text": "这是转写测试。"},
                    ],
                    {"language": "zh", "duration": 3.0},
                )

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "speech.mp3"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_bytes(b"fake mp3 payload")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="audio",
                origin_name="speech.mp3",
                storage_relpath="master/speech.mp3",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="speech.mp3",
                short_hint="一段口播录音。",
                detail={"file_kind": "mp3", "media_info": {"audio": {"codec": "mp3"}}},
                timestamp=110,
            )

            def fake_run(command, capture_output, text, timeout, check):
                _write_test_wav(Path(command[-1]), sample_rate=1000)
                return subprocess.CompletedProcess(command, 0, "", "")

            with patch("companion_v01.generated_files.importlib.util.find_spec", return_value=object()), patch(
                "companion_v01.generated_files.shutil.which",
                return_value="ffmpeg.exe",
            ), patch("companion_v01.generated_files.subprocess.run", side_effect=fake_run), patch.object(
                GeneratedFileService,
                "_load_faster_whisper_model",
                return_value=FakeWhisperModel(),
            ):
                result = generated_service.transcribe_media(
                    profile_user_id="user",
                    session_id="session",
                    source_targets=["audio_001"],
                    output_format="md",
                    output_title="speech_transcript",
                    language="zh",
                    with_timestamps=True,
                    merge_outputs=True,
                    timestamp=120,
                )

            self.assertTrue(result["ok"])
            generated = result["generated"]
            self.assertEqual(generated["generated_handle"], "gen_001")
            self.assertEqual(generated["output_format"], "md")
            self.assertEqual(generated["created_by_tool"], "transcribe_media")
            output_text = Path(generated["absolute_path"]).read_text(encoding="utf-8")
            self.assertIn("你好，主人", output_text)
            self.assertIn("[0:00 - 0:01]", output_text)
            card = generated.get("content_card") or {}
            self.assertEqual(card.get("type"), "media_transcript")
            self.assertIn("转写稿", result["followup_context"])

    def test_transcribe_media_tool_handler_emits_multiple_generated_events(self) -> None:
        class FakeGeneratedService:
            def transcribe_media(self, **kwargs):
                return {
                    "ok": True,
                    "generated": {"generated_id": "generated::1", "generated_handle": "gen_001"},
                    "generated_files": [
                        {"generated_id": "generated::1", "generated_handle": "gen_001"},
                        {"generated_id": "generated::2", "generated_handle": "gen_002"},
                    ],
                    "send_to_user": True,
                    "followup_context": "转写完成。",
                }

        handler = TranscribeMediaToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call(
            {
                "type": "transcribe_media",
                "source_ids": ["audio_001", "gen_002"],
                "output_format": "srt",
                "language": "中文",
                "merge_outputs": False,
                "with_timestamps": True,
                "send_to_user": True,
            }
        )

        self.assertIsNotNone(call)
        self.assertEqual((call or {}).get("source_ids"), ["audio_001", "gen_002"])
        self.assertEqual((call or {}).get("output_format"), "srt")
        self.assertEqual((call or {}).get("language"), "zh")
        self.assertFalse((call or {}).get("merge_outputs"))
        result = handler.execute(
            call=call or {},
            context=ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=100,
                visual_payload={},
            ),
        )

        self.assertEqual(len(result.stream_events), 2)
        self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
        self.assertEqual(result.stream_events[1]["generated_file"]["generated_handle"], "gen_002")
        self.assertEqual(result.followup_context, "转写完成。")

    def test_inspect_media_info_tool_handler_emits_media_info_event(self) -> None:
        class FakeGeneratedService:
            def inspect_media_info(self, **kwargs):
                return {
                    "ok": True,
                    "media_info": {"duration_seconds": 12.0, "audio": {"codec": "mp3"}},
                    "followup_context": "读取完成。",
                }

        handler = InspectMediaInfoToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call({"type": "inspect_media_info", "source_id": "audio_001"})

        self.assertIsNotNone(call)
        result = handler.execute(
            call=call or {},
            context=ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=100,
                visual_payload={},
            ),
        )

        self.assertEqual(result.stream_events[0]["type"], "media_info_inspected")
        self.assertEqual(result.stream_events[0]["media_info"]["duration_seconds"], 12.0)
        self.assertEqual(result.followup_context, "读取完成。")

    def test_convert_media_file_rejects_protected_platform_formats(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "song.kgm"
            stored.parent.mkdir(parents=True, exist_ok=True)
            stored.write_bytes(b"protected payload")

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="audio",
                origin_name="song.kgm",
                storage_relpath="master/song.kgm",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="song.kgm",
                short_hint="一首平台缓存音频。",
                detail={"file_kind": "kgm"},
                timestamp=110,
            )

            result = generated_service.convert_media_file(
                profile_user_id="user",
                session_id="session",
                source_target="audio_001",
                output_format="mp3",
                output_title="song_mp3",
                timestamp=120,
            )

            self.assertFalse(result["ok"])
            self.assertEqual(result["error"], "protected_media_format")
            self.assertIn("不要尝试解密", result["followup_context"])

    def test_convert_media_file_tool_handler_emits_generated_event(self) -> None:
        class FakeGeneratedService:
            def convert_media_file(self, **kwargs):
                return {
                    "ok": True,
                    "generated": {"generated_id": "generated::1", "generated_handle": "gen_001"},
                    "send_to_user": True,
                    "followup_context": "转换完成。",
                }

        handler = ConvertMediaFileToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call(
            {
                "type": "convert_media_file",
                "source_id": "audio_001",
                "output_format": "wav",
                "start_time": "1:20",
                "end_time": "2:00",
                "normalize_volume": True,
                "volume_gain_db": "6db",
                "trim_silence": True,
                "fade_in_seconds": 1.5,
                "fade_out_seconds": 2,
                "speed_ratio": "125%",
                "send_to_user": True,
            }
        )

        self.assertIsNotNone(call)
        self.assertEqual((call or {}).get("start_time"), "1:20")
        self.assertEqual((call or {}).get("volume_gain_db"), 6.0)
        self.assertTrue((call or {}).get("trim_silence"))
        self.assertEqual((call or {}).get("speed_ratio"), 1.25)
        result = handler.execute(
            call=call or {},
            context=ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=100,
                visual_payload={},
            ),
        )

        self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
        self.assertEqual(result.followup_context, "转换完成。")

    def test_convert_media_file_tool_instruction_keeps_audio_options_optional(self) -> None:
        handler = ConvertMediaFileToolHandler(generated_file_service=object())

        instruction = handler.build_prompt_instruction()

        self.assertIn("视频", instruction)
        self.assertIn("sample_rate", instruction)
        self.assertIn("可选项", instruction)
        self.assertIn("不要硬填", instruction)
        self.assertIn("16000", instruction)
        self.assertIn("start_time", instruction)
        self.assertIn("volume_gain_db", instruction)
        self.assertIn("trim_silence", instruction)
        self.assertIn("fade_out_seconds", instruction)
        self.assertIn("speed_ratio", instruction)
        self.assertIn("只有用户要音频轨、后续人声处理、训练素材或统一媒体规格时才提音频", instruction)
        self.assertIn("如果用户只要原视频，改用 send_file 发送原文件", instruction)

    def test_media_tool_instructions_explain_video_task_routing_without_fixed_pipeline(self) -> None:
        transcribe_instruction = TranscribeMediaToolHandler(generated_file_service=object()).build_prompt_instruction()
        dataset_instruction = PrepareVoiceDatasetToolHandler(generated_file_service=object()).build_prompt_instruction()

        self.assertIn("如果用户要字幕文件，优先用 srt 或 vtt", transcribe_instruction)
        self.assertIn("这个工具负责转写，不负责总结", transcribe_instruction)
        self.assertIn("如果用户只要原视频/原音频，不要为了回复而转写", transcribe_instruction)
        self.assertIn("训练素材任务可以分多步组合", dataset_instruction)
        self.assertIn("先 convert_media_file 提音频", dataset_instruction)
        self.assertIn("再 separate_audio_stems 拿人声", dataset_instruction)
        self.assertIn("再 clean_voice_track 降噪", dataset_instruction)
        self.assertIn("最后 prepare_voice_dataset 切片打包", dataset_instruction)
        self.assertIn("不要把这些步骤用于只要原文件的请求", dataset_instruction)

    def test_compose_file_applies_xlsx_formatting(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )

            result = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="生成带样式的成绩表",
                output_format="xlsx",
                output_title="成绩表",
                table_rows=[
                    ["姓名", "评价"],
                    ["Akane", "合格"],
                    ["Miku", "优秀"],
                ],
                formatting={
                    "header": {"bold": True},
                    "columns": [{"match_header": "姓名", "font_color": "red"}],
                    "rows": [{"index": 2, "fill_color": "yellow"}],
                    "highlights": [{"text": "优秀", "fill_color": "orange", "bold": True}],
                },
                send_to_user=False,
                timestamp=140,
            )

            self.assertTrue(result["ok"])
            from openpyxl import load_workbook  # type: ignore

            workbook = load_workbook(Path(result["generated"]["absolute_path"]))
            sheet = workbook.active
            self.assertTrue(sheet["A1"].font.bold)
            self.assertTrue(str(sheet["A2"].font.color.rgb).endswith("FF0000"))
            self.assertTrue(str(sheet["A2"].fill.fgColor.rgb).endswith("FFFF00"))
            self.assertTrue(sheet["B3"].font.bold)
            self.assertTrue(str(sheet["B3"].fill.fgColor.rgb).endswith("FFC000"))
            self.assertEqual(result["generated"]["content_card"]["formatting"]["columns"][0]["font_color"], "FF0000")

    def test_revise_generated_file_creates_versioned_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            original = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份草稿",
                output_format="md",
                output_title="草稿",
                content_markdown="# 草稿\n\n第一段\n\n第二段",
                send_to_user=False,
                timestamp=100,
            )["generated"]

            revised = generated_service.revise_generated_file(
                profile_user_id="user",
                session_id="session",
                target="gen_001",
                instruction="删掉第二段，加一句总结",
                output_format="md",
                output_title="草稿修改版",
                content_markdown="# 草稿修改版\n\n第一段\n\n总结：已经更简洁。",
                timestamp=120,
            )

            self.assertTrue(revised["ok"])
            generated = revised["generated"]
            self.assertEqual(generated["generated_handle"], "gen_002")
            self.assertEqual(generated["version_of_generated_id"], original["generated_id"])
            self.assertEqual(generated["version_no"], 2)
            output_path = Path(generated["absolute_path"])
            self.assertIn("总结", output_path.read_text(encoding="utf-8"))

    def test_revise_generated_file_tool_handler_emits_generated_event(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="旧内容",
                timestamp=100,
            )
            handler = ReviseGeneratedFileToolHandler(generated_file_service=generated_service)
            call = handler.normalize_call(
                {
                    "type": "revise_generated_file",
                    "target": "gen_001",
                    "instruction": "改得更自然",
                    "content_markdown": "新内容",
                    "send_to_user": True,
                }
            )
            context = ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=130,
                visual_payload={},
            )

            result = handler.execute(call=call or {}, context=context)

            self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
            self.assertIn("gen_002", result.followup_context)
            self.assertTrue(result.stream_events[0]["send_to_user"])

    def test_send_generated_file_reuses_existing_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="旧内容",
                send_to_user=False,
                timestamp=100,
            )

            result = generated_service.send_generated_file(
                profile_user_id="user",
                session_id="session",
                target="gen_001",
                timestamp=130,
            )

            self.assertTrue(result["ok"])
            self.assertTrue(result["send_to_user"])
            self.assertEqual(result["generated"]["generated_handle"], "gen_001")
            self.assertEqual(result["generated"]["delivery_status"], "pending")
            self.assertTrue(Path(result["generated"]["absolute_path"]).exists())

    def test_send_generated_file_supports_multiple_targets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写第一份小结",
                output_format="txt",
                output_title="小结一",
                content_markdown="第一份内容",
                send_to_user=False,
                timestamp=100,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写第二份小结",
                output_format="txt",
                output_title="小结二",
                content_markdown="第二份内容",
                send_to_user=False,
                timestamp=110,
            )

            result = generated_service.send_generated_file(
                profile_user_id="user",
                session_id="session",
                targets=["gen_001", "gen_002"],
                timestamp=130,
            )

            self.assertTrue(result["ok"])
            self.assertEqual(len(result["generated_files"]), 2)
            self.assertEqual(
                [item["generated_handle"] for item in result["generated_files"]],
                ["gen_001", "gen_002"],
            )
            self.assertIn("2 份生成文件", result["followup_context"])

    def test_send_generated_file_tool_handler_emits_generated_event(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="旧内容",
                send_to_user=False,
                timestamp=100,
            )
            handler = SendGeneratedFileToolHandler(generated_file_service=generated_service)
            call = handler.normalize_call(
                {
                    "type": "send_generated_file",
                    "target": "gen_001",
                }
            )
            context = ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=130,
                visual_payload={},
            )

            result = handler.execute(call=call or {}, context=context)

            self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
            self.assertEqual(result.stream_events[0]["generated_file"]["generated_handle"], "gen_001")
            self.assertTrue(result.stream_events[0]["send_to_user"])
            self.assertIn("send_generated_file", result.followup_context)

    def test_send_generated_file_tool_handler_emits_multiple_generated_events(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写第一份小结",
                output_format="txt",
                output_title="小结一",
                content_markdown="第一份内容",
                send_to_user=False,
                timestamp=100,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写第二份小结",
                output_format="txt",
                output_title="小结二",
                content_markdown="第二份内容",
                send_to_user=False,
                timestamp=110,
            )
            handler = SendGeneratedFileToolHandler(generated_file_service=generated_service)
            call = handler.normalize_call(
                {
                    "type": "send_generated_file",
                    "targets": ["gen_001", "gen_002"],
                }
            )
            context = ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=130,
                visual_payload={},
            )

            result = handler.execute(call=call or {}, context=context)

            self.assertEqual(len(result.stream_events), 2)
            self.assertEqual(
                [event["generated_file"]["generated_handle"] for event in result.stream_events],
                ["gen_001", "gen_002"],
            )
            self.assertTrue(all(event["send_to_user"] for event in result.stream_events))

    def test_send_file_supports_generated_and_attachment_targets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root / "db")
            attachment_root = root / "attachments"
            attachment_path = attachment_root / "master" / "clip.mp4"
            attachment_path.parent.mkdir(parents=True, exist_ok=True)
            attachment_path.write_bytes(b"video")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="video",
                origin_name="clip.mp4",
                storage_relpath="master/clip.mp4",
                file_ext="mp4",
                timestamp=90,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="测试视频",
                short_hint="一个测试视频。",
                detail={"media_info": {"video": {"codec": "h264"}}},
                timestamp=95,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="生成内容",
                send_to_user=False,
                timestamp=100,
            )

            result = generated_service.send_file(
                profile_user_id="user",
                session_id="session",
                targets=["file_001", "gen_001"],
                timestamp=130,
            )

            self.assertTrue(result["ok"])
            self.assertEqual([item["source_type"] for item in result["files"]], ["attachment", "generated"])
            self.assertEqual(result["files"][0]["name"], "clip.mp4")
            self.assertEqual(result["files"][1]["handle"], "gen_001")
            self.assertIn("2 个已有文件", result["followup_context"])

    def test_send_file_requests_confirmation_for_ambiguous_generated_name(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            first = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份原始人声说明",
                output_format="txt",
                output_title="昔涟_人声",
                content_markdown="原始人声文件",
                send_to_user=False,
                timestamp=100,
            )["generated"]
            second = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份降噪人声说明",
                output_format="txt",
                output_title="昔涟_人声_降噪",
                content_markdown="降噪人声文件",
                send_to_user=False,
                timestamp=110,
            )["generated"]

            result = generated_service.send_file(
                profile_user_id="user",
                session_id="session",
                targets=["人声"],
                timestamp=130,
            )

            self.assertFalse(result["ok"])
            self.assertEqual(result["error"], "file_target_ambiguous")
            self.assertIn("存在多个候选", result["followup_context"])
            self.assertIn(first["generated_handle"], result["followup_context"])
            self.assertIn(second["generated_handle"], result["followup_context"])

    def test_send_file_keeps_generated_exact_send_when_attachment_target_is_ambiguous(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root / "db")
            attachment_root = root / "attachments"
            first_path = attachment_root / "master" / "menu-breakfast.png"
            second_path = attachment_root / "master" / "menu-dinner.png"
            first_path.parent.mkdir(parents=True, exist_ok=True)
            first_path.write_bytes(b"img1")
            second_path.write_bytes(b"img2")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            first = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="image",
                origin_name="menu-breakfast.png",
                storage_relpath="master/menu-breakfast.png",
                timestamp=90,
            )
            second = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="image",
                origin_name="menu-dinner.png",
                storage_relpath="master/menu-dinner.png",
                timestamp=91,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=first["attachment_id"],
                summary_title="早餐菜单图",
                short_hint="第一张菜单图。",
                timestamp=95,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=second["attachment_id"],
                summary_title="晚餐菜单图",
                short_hint="第二张菜单图。",
                timestamp=96,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="生成内容",
                send_to_user=False,
                timestamp=100,
            )

            result = generated_service.send_file(
                profile_user_id="user",
                session_id="session",
                targets=["菜单图", "gen_001"],
                timestamp=130,
            )

            self.assertTrue(result["ok"])
            self.assertEqual([item["handle"] for item in result["files"]], ["gen_001"])
            self.assertIn("存在多个候选", result["followup_context"])
            self.assertIn("img_001", result["followup_context"])
            self.assertIn("img_002", result["followup_context"])
            self.assertIn("gen_001", result["followup_context"])

    def test_send_file_tool_handler_emits_generic_file_events(self) -> None:
        class FakeGeneratedService:
            def send_file(self, **kwargs):
                return {
                    "ok": True,
                    "files": [
                        {
                            "source_type": "attachment",
                            "source_id": "attachment::1",
                            "handle": "file_001",
                            "absolute_path": "C:/tmp/video.mp4",
                            "name": "video.mp4",
                        },
                        {
                            "source_type": "generated",
                            "source_id": "generated::1",
                            "generated_id": "generated::1",
                            "handle": "gen_001",
                            "absolute_path": "C:/tmp/out.md",
                            "name": "out.md",
                        },
                    ],
                    "followup_context": "send_file ok",
                }

        handler = SendFileToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call(
            {
                "type": "send_file",
                "targets": ["file_001", "gen_001"],
            }
        )
        context = ToolExecutionContext(
            profile_user_id="user",
            session_id="session",
            now_ts=130,
            visual_payload={},
        )

        result = handler.execute(call=call or {}, context=context)

        self.assertEqual([event["type"] for event in result.stream_events], ["file_ready", "file_ready"])
        self.assertEqual(result.stream_events[0]["file"]["handle"], "file_001")
        self.assertEqual(result.stream_events[1]["file"]["generated_id"], "generated::1")
        self.assertIn("send_file ok", result.followup_context)

    def test_send_file_tool_handler_carries_desktop_delivery_action(self) -> None:
        class FakeGeneratedService:
            def send_file(self, **kwargs):
                return {
                    "ok": True,
                    "files": [
                        {
                            "source_type": "generated",
                            "source_id": "generated::1",
                            "generated_id": "generated::1",
                            "handle": "gen_001",
                            "absolute_path": "C:/tmp/out.txt",
                            "name": "out.txt",
                        }
                    ],
                    "followup_context": "send_file ok",
                }

        handler = SendFileToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call(
            {
                "type": "send_file",
                "target": "gen_001",
                "delivery_action": "save_desktop",
            }
        )
        context = ToolExecutionContext(
            profile_user_id="user",
            session_id="session",
            now_ts=130,
            visual_payload={},
            client_mode="desktop_pet",
        )

        result = handler.execute(call=call or {}, context=context)

        self.assertEqual(result.stream_events[0]["delivery_action"], "save_desktop")
        self.assertEqual(result.stream_events[0]["client_mode"], "desktop_pet")
        self.assertEqual(result.stream_events[0]["desktop_delivery"]["action"], "save_desktop")
        self.assertEqual(result.stream_events[0]["desktop_delivery"]["handle"], "gen_001")

    def test_send_file_tool_handler_ignores_desktop_delivery_action_for_qq(self) -> None:
        class FakeGeneratedService:
            def send_file(self, **kwargs):
                return {
                    "ok": True,
                    "files": [
                        {
                            "source_type": "generated",
                            "source_id": "generated::1",
                            "generated_id": "generated::1",
                            "handle": "gen_001",
                            "absolute_path": "C:/tmp/out.txt",
                            "name": "out.txt",
                        }
                    ],
                    "followup_context": "send_file ok",
                }

        handler = SendFileToolHandler(generated_file_service=FakeGeneratedService())
        call = handler.normalize_call(
            {
                "type": "send_file",
                "target": "gen_001",
                "delivery_action": "save_desktop",
            }
        )
        context = ToolExecutionContext(
            profile_user_id="user",
            session_id="session",
            now_ts=130,
            visual_payload={},
            client_mode="qq_text",
        )

        result = handler.execute(call=call or {}, context=context)

        self.assertEqual(result.stream_events[0]["client_mode"], "qq_text")
        self.assertNotIn("delivery_action", result.stream_events[0])
        self.assertNotIn("desktop_delivery", result.stream_events[0])

    def test_inspect_generated_file_reads_text_content(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="md",
                output_title="小结",
                content_markdown="# 小结\n\n第一段内容。\n\n最后一段内容。",
                send_to_user=False,
                timestamp=100,
            )

            result = generated_service.inspect_generated_file(
                profile_user_id="user",
                session_id="session",
                target="gen_001",
                section="tail",
                max_chars=40,
            )

            self.assertTrue(result["ok"])
            self.assertEqual(result["inspection"]["section"], "tail")
            self.assertIn("最后一段内容", result["inspection"]["content"])
            self.assertIn("inspect_generated_file", result["followup_context"])

    def test_inspect_generated_file_reads_zip_manifest_and_list(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            zip_path = root / "generated_files" / "user" / "session" / "dataset.zip"
            zip_path.parent.mkdir(parents=True, exist_ok=True)
            manifest = {"title": "测试训练集", "stats": {"slice_count": 2}}
            with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
                archive.writestr("README.md", "# 测试训练集\n")
                archive.writestr("slices/clip_001.wav", b"RIFF")
            store.add_generated_file(
                profile_user_id="user",
                session_id="session",
                output_title="测试训练集",
                output_format="zip",
                storage_relpath="user/session/dataset.zip",
                mime_type="application/zip",
                file_ext="zip",
                file_size=zip_path.stat().st_size,
                content_card={"summary": "训练集 zip"},
                summary="训练集 zip",
                timestamp=100,
            )

            manifest_result = generated_service.inspect_generated_file(
                profile_user_id="user",
                session_id="session",
                target="gen_001",
                section="manifest",
                max_chars=2000,
            )
            list_result = generated_service.inspect_generated_file(
                profile_user_id="user",
                session_id="session",
                target="gen_001",
                section="file_list",
                max_chars=2000,
            )

            self.assertTrue(manifest_result["ok"])
            self.assertIn("slice_count", manifest_result["inspection"]["content"])
            self.assertIn("manifest.json", list_result["inspection"]["content"])
            self.assertIn("slices/clip_001.wav", list_result["inspection"]["content"])

    def test_inspect_generated_file_tool_handler_emits_inspection_event(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="可以回头查看的内容",
                send_to_user=False,
                timestamp=100,
            )
            handler = InspectGeneratedFileToolHandler(generated_file_service=generated_service)
            call = handler.normalize_call(
                {
                    "type": "inspect_generated_file",
                    "target": "gen_001",
                    "section": "content",
                }
            )
            context = ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=130,
                visual_payload={},
            )

            result = handler.execute(call=call or {}, context=context)

            self.assertEqual(result.stream_events[0]["type"], "generated_file_inspected")
            self.assertIn("可以回头查看", result.followup_context)

    def test_apply_style_to_existing_attachment_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            attachment_root = root / "attachments"
            stored = attachment_root / "master" / "scores.xlsx"
            stored.parent.mkdir(parents=True, exist_ok=True)

            from openpyxl import Workbook, load_workbook  # type: ignore

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["姓名", "分数"])
            sheet.append(["Akane", 95])
            sheet.append(["Miku", 55])
            workbook.save(stored)

            store = MemoryStore(root / "db")
            attachment_service = AttachmentInboxService(store=store, base_dir=attachment_root)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            attachment = attachment_service.create_pending(
                profile_user_id="user",
                session_id="session",
                source="qq",
                kind="document",
                origin_name="scores.xlsx",
                storage_relpath="master/scores.xlsx",
                timestamp=100,
            )
            attachment_service.mark_ready(
                profile_user_id="user",
                session_id="session",
                attachment_id=attachment["attachment_id"],
                summary_title="成绩表",
                short_hint="包含姓名和分数。",
                detail={"file_kind": "xlsx", "sheets": [{"name": "Sheet", "rows": [["姓名", "分数"]]}]},
                timestamp=110,
            )

            result = generated_service.apply_style_to_existing_file(
                profile_user_id="user",
                session_id="session",
                target="file_001",
                target_type="attachment",
                instruction="低于60分整行标红，姓名列加粗",
                output_title="成绩表标注版",
                formatting={
                    "columns": [{"match_header": "姓名", "bold": True}],
                    "row_rules": [{"where": {"column": "分数", "lt": 60}, "font_color": "red"}],
                },
                send_to_user=False,
                timestamp=130,
            )

            self.assertTrue(result["ok"])
            generated = result["generated"]
            self.assertEqual(generated["generated_handle"], "gen_001")
            self.assertEqual(generated["created_by_tool"], "apply_style_to_existing_file")
            output_path = Path(generated["absolute_path"])
            self.assertTrue(output_path.exists())
            styled = load_workbook(output_path)
            styled_sheet = styled.active
            self.assertTrue(styled_sheet["A2"].font.bold)
            self.assertTrue(styled_sheet["A3"].font.bold)
            self.assertTrue(str(styled_sheet["A3"].font.color.rgb).endswith("FF0000"))
            self.assertTrue(str(styled_sheet["B3"].font.color.rgb).endswith("FF0000"))

    def test_apply_style_to_existing_file_tool_handler_emits_generated_event(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="生成表格",
                output_format="xlsx",
                output_title="成绩表",
                table_rows=[["姓名", "分数"], ["Akane", 95]],
                send_to_user=False,
                timestamp=100,
            )
            handler = ApplyStyleToExistingFileToolHandler(generated_file_service=generated_service)
            call = handler.normalize_call(
                {
                    "type": "apply_style_to_existing_file",
                    "target": "gen_001",
                    "target_type": "generated",
                    "instruction": "姓名列标红",
                    "formatting": {"columns": [{"match_header": "姓名", "font_color": "red"}]},
                    "send_to_user": True,
                }
            )
            context = ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=130,
                visual_payload={},
            )

            result = handler.execute(call=call or {}, context=context)

            self.assertEqual(result.stream_events[0]["type"], "generated_file_ready")
            self.assertEqual(result.stream_events[0]["generated_file"]["generated_handle"], "gen_002")
            self.assertTrue(result.stream_events[0]["send_to_user"])
            self.assertIn("apply_style_to_existing_file", result.followup_context)

    def test_manage_generated_files_archive_hides_from_prompt(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="旧内容",
                send_to_user=False,
                timestamp=100,
            )

            result = generated_service.manage_generated_files(
                profile_user_id="user",
                session_id="session",
                action="archive",
                targets=["gen_001"],
                reason="旧版本不用了",
                timestamp=130,
            )

            self.assertTrue(result["ok"])
            self.assertEqual(result["managed"][0]["status"], "removed")
            self.assertEqual(
                generated_service.build_prompt_context(profile_user_id="user", session_id="session"),
                "",
            )

    def test_manage_generated_files_delete_removes_disk_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated = generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="旧内容",
                send_to_user=False,
                timestamp=100,
            )["generated"]
            output_path = Path(generated["absolute_path"])
            self.assertTrue(output_path.exists())

            result = generated_service.manage_generated_files(
                profile_user_id="user",
                session_id="session",
                action="delete",
                targets=["gen_001"],
                timestamp=130,
            )

            self.assertTrue(result["ok"])
            self.assertFalse(output_path.exists())
            self.assertTrue(result["managed"][0]["file_deleted"])
            self.assertEqual(result["managed"][0]["status"], "removed")

    def test_manage_generated_file_tool_handler_emits_event(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            store = MemoryStore(root)
            attachment_service = AttachmentInboxService(store=store)
            generated_service = GeneratedFileService(
                base_dir=root / "generated_files",
                store=store,
                attachment_service=attachment_service,
            )
            generated_service.compose_file(
                profile_user_id="user",
                session_id="session",
                source_targets=[],
                task="写一份小结",
                output_format="txt",
                output_title="小结",
                content_markdown="旧内容",
                send_to_user=False,
                timestamp=100,
            )
            handler = ManageGeneratedFileToolHandler(generated_file_service=generated_service)
            call = handler.normalize_call(
                {
                    "type": "manage_generated_file",
                    "action": "archive",
                    "targets": ["gen_001"],
                    "reason": "测试清理",
                }
            )
            context = ToolExecutionContext(
                profile_user_id="user",
                session_id="session",
                now_ts=130,
                visual_payload={},
            )

            result = handler.execute(call=call or {}, context=context)

            self.assertEqual(result.stream_events[0]["type"], "generated_files_managed")
            self.assertEqual(result.stream_events[0]["action"], "archive")
            self.assertIn("manage_generated_file", result.followup_context)


if __name__ == "__main__":
    unittest.main()
